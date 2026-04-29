"""
xlsx_converter.py - XLSX → Markdown 转换器

读取 Excel 文件，将每个 sheet 转为 markdown 表格。
不依赖 MinerU，直接用 openpyxl 读取。
"""

import os
from typing import Optional


def convert_xlsx_to_markdown(xlsx_path: str, output_dir: str, skip_if_exists: bool = True) -> dict:
    """
    将 XLSX 文件转换为 Markdown 表格

    Args:
        xlsx_path: XLSX 文件路径
        output_dir: 输出目录
        skip_if_exists: 如果 MD 已存在则跳过

    Returns:
        {
            "success": bool,
            "markdown_path": str,
            "markdown_text": str,
            "skipped": bool,
            "error": str,
        }
    """
    if not os.path.exists(xlsx_path):
        return {"success": False, "error": f"文件不存在: {xlsx_path}"}

    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    md_path = os.path.join(output_dir, f"{base}.md")

    if skip_if_exists and os.path.exists(md_path):
        return {"success": True, "markdown_path": md_path, "skipped": True}

    try:
        import openpyxl
    except ImportError:
        return {"success": False, "error": "openpyxl 未安装: pip install openpyxl"}

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        return {"success": False, "error": f"无法打开 XLSX: {e}"}

    sheets_data = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            # 跳过全空行
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            rows.append([str(cell) if cell is not None else "" for cell in row])

        if not rows:
            continue

        sheets_data.append((sheet_name, rows))

    wb.close()

    if not sheets_data:
        return {"success": False, "error": "XLSX 文件为空或无有效数据"}

    # 生成 Markdown
    md_parts = []
    for sheet_name, rows in sheets_data:
        if len(sheets_data) > 1:
            md_parts.append(f"## {sheet_name}\n")

        # 表头
        header = rows[0]
        md_parts.append("| " + " | ".join(header) + " |")
        md_parts.append("| " + " | ".join(["---"] * len(header)) + " |")

        # 数据行
        for row in rows[1:]:
            # 对齐列数
            padded = row + [""] * (len(header) - len(row))
            md_parts.append("| " + " | ".join(padded[:len(header)]) + " |")

        md_parts.append("")  # 空行分隔

    md_text = "\n".join(md_parts)

    os.makedirs(output_dir, exist_ok=True)
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md_text)

    return {
        "success": True,
        "markdown_path": md_path,
        "markdown_text": md_text,
        "skipped": False,
    }
